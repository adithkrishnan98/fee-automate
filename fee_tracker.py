#!/usr/bin/env python3
"""
Music Class Fee Tracker
A GUI application to categorize and track student fees from bank statements
"""

import sys
import re
import json
import copy
from datetime import datetime
from pathlib import Path
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QFileDialog,
    QLabel, QGroupBox, QHeaderView, QMessageBox, QLineEdit, QComboBox,
    QDialog, QFormLayout, QDialogButtonBox, QDoubleSpinBox, QGridLayout,
    QCheckBox, QDateEdit, QListWidget, QListWidgetItem
)
from PySide6.QtCore import Qt, QTimer, QDate
from PySide6.QtGui import QFont, QColor

# Import student name mappings
from student_mappings import get_full_name


class EditTransactionDialog(QDialog):
    """Dialog for editing a transaction"""
    
    def __init__(self, transaction, available_categories, parent=None):
        super().__init__(parent)
        self.transaction = transaction
        self.available_categories = available_categories
        self.setWindowTitle("Edit Transaction")
        self.setModal(True)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the dialog UI"""
        layout = QFormLayout(self)
        
        # Name field (editable)
        self.name_input = QLineEdit(self.transaction['name'])
        layout.addRow("Name:", self.name_input)
        
        # Date field (editable)
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd-MM-yyyy")
        
        # Parse the date from the transaction (format can be DD-MM-YYYY HH:MM:SS or DD-MM-YYYY or DD/MM/YYYY)
        date_str = self.transaction['date']
        # Remove time component if present
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]
        
        # Try splitting by hyphen first (the actual format from CSV)
        date_parts = date_str.split('-')
        if len(date_parts) != 3:
            # Try splitting by slash as fallback
            date_parts = date_str.split('/')
        
        if len(date_parts) == 3:
            try:
                day, month, year = date_parts
                self.date_edit.setDate(QDate(int(year), int(month), int(day)))
            except ValueError:
                # If parsing fails, use current date
                self.date_edit.setDate(QDate.currentDate())
        else:
            # If format is unexpected, use current date
            self.date_edit.setDate(QDate.currentDate())
        
        layout.addRow("Date:", self.date_edit)
        
        # Amount field (editable)
        self.amount_spinbox = QDoubleSpinBox()
        self.amount_spinbox.setRange(0, 999999)
        self.amount_spinbox.setDecimals(2)
        self.amount_spinbox.setValue(self.transaction['amount'])
        self.amount_spinbox.setPrefix("‚Çπ")
        layout.addRow("Amount:", self.amount_spinbox)
        
        # Category field (editable)
        self.category_combo = QComboBox()
        self.category_combo.addItems(self.available_categories)
        self.category_combo.setCurrentText(self.transaction['category'])
        layout.addRow("Category:", self.category_combo)
        
        # Dialog buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def get_updated_transaction(self):
        """Return the updated transaction data"""
        # Format date back to DD-MM-YYYY (matching the original CSV format)
        # Preserve the time component if it exists in the original
        date = self.date_edit.date()
        formatted_date = date.toString("dd-MM-yyyy")
        
        # If original date had time component, append it back
        original_date = self.transaction['date']
        if ' ' in original_date:
            time_component = ' '.join(original_date.split(' ')[1:])
            formatted_date = f"{formatted_date} {time_component}"
        
        return {
            **self.transaction,
            'name': self.name_input.text(),
            'date': formatted_date,
            'amount': self.amount_spinbox.value(),
            'category': self.category_combo.currentText()
        }


class CategoryManagerDialog(QDialog):
    """Dialog for managing fee categories"""
    
    def __init__(self, categories, parent=None):
        super().__init__(parent)
        self.categories = copy.deepcopy(categories)  # Deep copy to avoid reference issues
        self.setWindowTitle("Manage Categories")
        self.setModal(True)
        self.setMinimumSize(600, 400)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the category manager UI"""
        layout = QVBoxLayout(self)
        
        # Title
        title_label = QLabel("üìÇ Category Management")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Instructions
        info_label = QLabel("Add, edit, or delete fee categories. Categories with fixed amounts will auto-categorize transactions.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666; padding: 10px; background-color: #f5f5f5; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # Main content area
        content_layout = QHBoxLayout()
        
        # Left side - Category list
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("Existing Categories:"))
        
        self.category_list = QListWidget()
        self.category_list.setMinimumWidth(300)
        self.update_category_list()
        self.category_list.itemSelectionChanged.connect(self.on_category_selected)
        left_layout.addWidget(self.category_list)
        
        # List buttons
        list_btn_layout = QHBoxLayout()
        
        self.delete_btn = QPushButton("üóëÔ∏è Delete")
        self.delete_btn.setEnabled(False)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover:enabled {
                background-color: #d32f2f;
            }
            QPushButton:disabled {
                background-color: #ccc;
            }
        """)
        self.delete_btn.clicked.connect(self.delete_category)
        list_btn_layout.addWidget(self.delete_btn)
        
        list_btn_layout.addStretch()
        left_layout.addLayout(list_btn_layout)
        content_layout.addLayout(left_layout)
        
        # Right side - Add/Edit form
        right_layout = QVBoxLayout()
        form_label = QLabel("Add/Edit Category:")
        form_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #333;")
        right_layout.addWidget(form_label)
        
        form_widget = QWidget()
        form_widget.setStyleSheet("background-color: #fafafa; border: 1px solid #ddd; border-radius: 5px;")
        form_layout = QFormLayout(form_widget)
        
        # Category name
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("e.g., Piano Lessons")
        self.name_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: #000;
                padding: 8px;
                border: 2px solid #ccc;
                border-radius: 4px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border: 2px solid #2196F3;
            }
        """)
        
        name_label = QLabel("Category Name:")
        name_label.setStyleSheet("color: #333; font-weight: bold;")
        form_layout.addRow(name_label, self.name_input)
        
        # Fee amount
        self.fee_input = QDoubleSpinBox()
        self.fee_input.setRange(0, 999999)
        self.fee_input.setDecimals(2)
        self.fee_input.setValue(500.0)
        self.fee_input.setPrefix("‚Çπ")
        self.fee_input.setSpecialValueText("Variable Amount")
        self.fee_input.setStyleSheet("""
            QDoubleSpinBox {
                background-color: white;
                color: #000;
                padding: 8px;
                border: 2px solid #ccc;
                border-radius: 4px;
                font-size: 13px;
            }
            QDoubleSpinBox:focus {
                border: 2px solid #2196F3;
            }
        """)
        
        fee_label = QLabel("Fixed Amount:")
        fee_label.setStyleSheet("color: #333; font-weight: bold;")
        form_layout.addRow(fee_label, self.fee_input)
        
        # Help text
        help_text = QLabel("üí° Set amount to 0 for variable amounts (donations, etc.)")
        help_text.setStyleSheet("color: #666; font-size: 11px; padding: 5px;")
        form_layout.addRow("", help_text)
        
        right_layout.addWidget(form_widget)
        
        # Form buttons
        form_btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("‚ûï Add Category")
        self.add_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.add_btn.clicked.connect(self.add_category)
        form_btn_layout.addWidget(self.add_btn)
        
        self.update_btn = QPushButton("‚úèÔ∏è Update Category")
        self.update_btn.setEnabled(False)
        self.update_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover:enabled {
                background-color: #F57C00;
            }
            QPushButton:disabled {
                background-color: #ccc;
            }
        """)
        self.update_btn.clicked.connect(self.update_category)
        form_btn_layout.addWidget(self.update_btn)
        
        self.clear_btn = QPushButton("üîÑ Clear")
        self.clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                padding: 8px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #616161;
            }
        """)
        self.clear_btn.clicked.connect(self.clear_form)
        form_btn_layout.addWidget(self.clear_btn)
        
        right_layout.addLayout(form_btn_layout)
        right_layout.addStretch()
        
        content_layout.addLayout(right_layout)
        layout.addLayout(content_layout)
        
        # Dialog buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.save_btn = QPushButton("üíæ Save")
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 30px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        self.save_btn.clicked.connect(self.accept)
        button_layout.addWidget(self.save_btn)
        
        cancel_btn = QPushButton("‚ùå Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                padding: 10px 30px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #616161;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def update_category_list(self):
        """Update the category list display"""
        self.category_list.clear()
        
        # Create font for category items
        item_font = QFont()
        item_font.setPointSize(14)
        
        for category in self.categories:
            if category['fee'] is None or category['fee'] == 0:
                text = f"üìä {category['name']} (Variable Amount)"
            else:
                text = f"üí∞ {category['name']} (‚Çπ{category['fee']:.0f})"
            
            item = QListWidgetItem(text)
            item.setFont(item_font)
            item.setData(Qt.UserRole, category)
            self.category_list.addItem(item)
    
    def on_category_selected(self):
        """Handle category selection"""
        current_item = self.category_list.currentItem()
        if current_item:
            category = current_item.data(Qt.UserRole)
            self.name_input.setText(category['name'])
            self.fee_input.setValue(category['fee'] if category['fee'] is not None else 0)
            
            self.delete_btn.setEnabled(True)
            self.update_btn.setEnabled(True)
            self.add_btn.setText("‚ûï Add New Category")
        else:
            self.delete_btn.setEnabled(False)
            self.update_btn.setEnabled(False)
    
    def add_category(self):
        """Add a new category"""
        name = self.name_input.text().strip()
        fee = self.fee_input.value()
        
        if not name:
            QMessageBox.warning(self, "Invalid Input", "Please enter a category name.")
            return
        
        # Check for duplicate names
        if any(cat['name'].lower() == name.lower() for cat in self.categories):
            QMessageBox.warning(self, "Duplicate Category", f"Category '{name}' already exists.")
            return
        
        # Check for duplicate variable amount categories (fee = 0 or null)
        if fee == 0:
            existing_variable = next((cat for cat in self.categories if cat['fee'] is None or cat['fee'] == 0), None)
            if existing_variable:
                QMessageBox.warning(
                    self, 
                    "Variable Amount Exists", 
                    f"A variable amount category '{existing_variable['name']}' already exists.\n\n"
                    f"You can only have one variable amount category for miscellaneous/donation transactions.\n"
                    f"Please use a fixed amount instead or edit the existing variable category."
                )
                return
        
        # Check for duplicate fee amounts (only for non-zero fees)
        if fee > 0:
            existing_category = next((cat for cat in self.categories if cat['fee'] == fee), None)
            if existing_category:
                QMessageBox.warning(
                    self, 
                    "Duplicate Amount", 
                    f"Amount ‚Çπ{fee:.0f} is already used by '{existing_category['name']}'.\n\n"
                    f"Each category must have a unique fixed amount for proper auto-categorization.\n"
                    f"Please use a different amount or set to 0 for variable amounts."
                )
                return
        
        # Add category
        new_category = {
            'name': name,
            'fee': fee if fee > 0 else None
        }
        self.categories.append(new_category)
        
        self.update_category_list()
        self.clear_form()
        
        QMessageBox.information(self, "Success", f"Category '{name}' added successfully!")
    
    def update_category(self):
        """Update selected category"""
        current_item = self.category_list.currentItem()
        if not current_item:
            return
        
        name = self.name_input.text().strip()
        fee = self.fee_input.value()
        
        if not name:
            QMessageBox.warning(self, "Invalid Input", "Please enter a category name.")
            return
        
        # Get selected category from the list item
        selected_category = current_item.data(Qt.UserRole)
        old_name = selected_category['name']
        
        # Find the actual category in self.categories list by matching the old values
        category_to_update = None
        for cat in self.categories:
            if cat['name'] == old_name and cat['fee'] == selected_category['fee']:
                category_to_update = cat
                break
        
        if not category_to_update:
            QMessageBox.warning(self, "Error", "Could not find category to update.")
            return
        
        # Check for duplicate names (excluding current)
        if any(cat['name'].lower() == name.lower() and cat is not category_to_update for cat in self.categories):
            QMessageBox.warning(self, "Duplicate Category", f"Category '{name}' already exists.")
            return
        
        # Check for duplicate variable amount categories (excluding current)
        if fee == 0:
            existing_variable = next((cat for cat in self.categories if (cat['fee'] is None or cat['fee'] == 0) and cat is not category_to_update), None)
            if existing_variable:
                QMessageBox.warning(
                    self, 
                    "Variable Amount Exists", 
                    f"A variable amount category '{existing_variable['name']}' already exists.\n\n"
                    f"You can only have one variable amount category for miscellaneous/donation transactions.\n"
                    f"Please use a fixed amount instead or edit the existing variable category."
                )
                return
        
        # Check for duplicate fee amounts (only for non-zero fees, excluding current)
        if fee > 0:
            existing_category = next((cat for cat in self.categories if cat['fee'] == fee and cat is not category_to_update), None)
            if existing_category:
                QMessageBox.warning(
                    self, 
                    "Duplicate Amount", 
                    f"Amount ‚Çπ{fee:.0f} is already used by '{existing_category['name']}'.\n\n"
                    f"Each category must have a unique fixed amount for proper auto-categorization.\n"
                    f"Please use a different amount or set to 0 for variable amounts."
                )
                return
        
        # Update the category in the list
        category_to_update['name'] = name
        category_to_update['fee'] = fee if fee > 0 else None
        
        # Refresh the list display
        self.update_category_list()
        self.clear_form()
        
        QMessageBox.information(self, "Success", f"Category updated from '{old_name}' to '{name}'!")
    
    def delete_category(self):
        """Delete selected category"""
        current_item = self.category_list.currentItem()
        if not current_item:
            return
        
        category = current_item.data(Qt.UserRole)
        
        reply = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete the category '{category['name']}'?\n\n"
            f"This action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.categories.remove(category)
            self.update_category_list()
            self.clear_form()
            
            QMessageBox.information(self, "Deleted", f"Category '{category['name']}' deleted successfully!")
    
    def clear_form(self):
        """Clear the form inputs"""
        self.name_input.clear()
        self.fee_input.setValue(500.0)
        self.category_list.clearSelection()
        
        self.delete_btn.setEnabled(False)
        self.update_btn.setEnabled(False)
        self.add_btn.setText("‚ûï Add Category")
    
    def get_categories(self):
        """Return the updated categories"""
        return self.categories


class FeeTrackerApp(QMainWindow):
    """Main application window for fee tracking"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Music Class Fee Tracker")
        self.setGeometry(100, 100, 1400, 800)
        
        # Load categories from JSON file
        self.categories = self.load_categories()
        
        # Data storage - build categorized_data dynamically
        self.transactions = []
        self.categorized_data = {}
        for cat in self.categories:
            category_key = self.get_category_key(cat['name'], cat['fee'])
            self.categorized_data[category_key] = []
        
        # Search optimization
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.perform_search)
        self.search_delay = 150  # milliseconds - faster response
        self.color_cache = {}  # Cache for color mappings
        
        # Selection state tracking
        self.selected_transactions = set()  # Track selected transaction IDs
        self._manual_selection_panel_open = False  # Track manual panel state
        self._updating_display = False  # Prevent concurrent updates
        
        # File tracking for auto-save
        self.current_csv_file = None  # Track currently loaded CSV
        self.auto_save_file = None  # Track auto-save JSON file
        
        self.init_ui()
    
    def load_categories(self):
        """Load categories from JSON configuration file"""
        self.config_file = Path(__file__).parent / 'fee_categories.json'
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data['categories']
        except FileNotFoundError:
            # Return default categories if file doesn't exist
            default_categories = [
                {"name": "Uncategorised", "fee": None}
            ]
            # Create the file with defaults
            self.save_categories(default_categories)
            return default_categories
        except Exception as e:
            QMessageBox.warning(self, "Warning", f"Error loading categories: {e}\nUsing defaults.")
            return [
                {"name": "Uncategorised", "fee": None}
            ]
    
    def save_categories(self, categories):
        """Save categories to JSON configuration file"""
        try:
            data = {"categories": categories}
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save categories: {e}")
            return False
    
    def get_category_key(self, name, fee):
        """Generate a consistent category key for storage"""
        if fee is not None:
            return f"{name} (‚Çπ{fee:.0f})"
        else:
            return name
    
    def get_transaction_id(self, transaction):
        """Generate a unique ID for a transaction"""
        return f"{transaction['date']}_{transaction['name']}_{transaction['amount']}_{transaction['reference']}"
    
    def get_auto_save_path(self, csv_file_path):
        """Generate path for auto-save JSON file"""
        csv_path = Path(csv_file_path)
        json_filename = csv_path.stem + '_edits.json'
        return csv_path.parent / json_filename
    
    def auto_save_data(self):
        """Auto-save current transactions to JSON file"""
        if not self.auto_save_file or not self.transactions:
            return
        
        try:
            save_data = {
                'csv_file': str(self.current_csv_file),
                'saved_at': datetime.now().isoformat(),
                'transactions': self.transactions,
                'categories': self.categories
            }
            
            with open(self.auto_save_file, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, indent=2, ensure_ascii=False)
            
            # Update status bar or label to show auto-saved
            self.file_label.setStyleSheet(
                "padding: 10px; background-color: #d4edda; border-radius: 5px; color: #155724;"
            )
            self.file_label.setText(f"File: {Path(self.current_csv_file).name} ‚úì Auto-saved")
            
            # Reset color after 2 seconds
            QTimer.singleShot(2000, lambda: self.file_label.setStyleSheet(
                "padding: 10px; background-color: #f0f0f0; border-radius: 5px; color: #212121;"
            ))
            
        except Exception as e:
            print(f"Auto-save failed: {e}")
    
    def load_saved_data(self, csv_file_path):
        """Load previously saved data if exists"""
        json_path = self.get_auto_save_path(csv_file_path)
        
        if not json_path.exists():
            return None
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                saved_data = json.load(f)
            
            # Verify it's for the same CSV file
            if saved_data.get('csv_file') == str(csv_file_path):
                return saved_data
            
        except Exception as e:
            print(f"Failed to load saved data: {e}")
        
        return None
        
    def init_ui(self):
        """Initialize the user interface"""
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title_label = QLabel("üéµ Music Class Fee Tracker")
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # File selection area
        file_layout = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("padding: 10px; background-color: #f0f0f0; border-radius: 5px; color: #212121;")
        file_layout.addWidget(self.file_label, 1)
        
        upload_btn = QPushButton("üìÅ Upload CSV File")
        upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        upload_btn.clicked.connect(self.upload_file)
        file_layout.addWidget(upload_btn)
        
        export_btn = QPushButton("üíæ Export Summary")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        export_btn.clicked.connect(self.export_summary)
        file_layout.addWidget(export_btn)
        
        # Category management button
        category_btn = QPushButton("‚öôÔ∏è Manage Categories")
        category_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        category_btn.clicked.connect(self.manage_categories)
        file_layout.addWidget(category_btn)
        
        main_layout.addLayout(file_layout)
        
        # Summary statistics
        self.summary_group = QGroupBox("Summary")
        self.summary_layout = QGridLayout()
        self.summary_layout.setSpacing(10)
        
        # Create dynamic summary labels based on categories
        self.category_labels = {}
        summary_font = QFont()
        summary_font.setPointSize(14)  # Increased from 11 to 14
        summary_font.setBold(True)
        
        # Calculate how many items per row (let's do 4 per row for good spacing)
        items_per_row = 4
        row = 0
        col = 0
        
        for cat in self.categories:
            category_key = self.get_category_key(cat['name'], cat['fee'])
            label = QLabel(f"{cat['name']}: 0 entries (‚Çπ0)")
            label.setFont(summary_font)
            label.setStyleSheet("padding: 12px; background-color: #e3f2fd; border-radius: 5px; color: #1a237e;")  # Increased padding
            self.category_labels[category_key] = label
            
            # Add to grid layout
            self.summary_layout.addWidget(label, row, col)
            col += 1
            if col >= items_per_row:
                col = 0
                row += 1
                if row >= 2:  # Maximum 2 rows
                    break
        
        self.summary_group.setLayout(self.summary_layout)
        main_layout.addWidget(self.summary_group)
        
        # Search toggle button
        self.search_toggle_btn = QPushButton("üîç Show Search")
        self.search_toggle_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)
        self.search_toggle_btn.clicked.connect(self.toggle_search)
        main_layout.addWidget(self.search_toggle_btn)
        
        # Search functionality (collapsible)
        self.search_group = QGroupBox("Search Transactions")
        search_layout = QHBoxLayout()
        
        # Search field dropdown
        search_label = QLabel("Search by:")
        search_layout.addWidget(search_label)
        
        self.search_field_combo = QComboBox()
        self.search_field_combo.addItems(["All", "Name", "Amount", "Category", "Description"])
        self.search_field_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 2px solid #2196F3;
                border-radius: 5px;
                min-width: 120px;
            }
        """)
        search_layout.addWidget(self.search_field_combo)
        
        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Enter search term...")
        self.search_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #2196F3;
                border-radius: 5px;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        self.search_input.textChanged.connect(self.on_search_text_changed)
        search_layout.addWidget(self.search_input, 1)
        
        # Clear search button
        clear_btn = QPushButton("‚úï Clear")
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        clear_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(clear_btn)
        
        # Search results label
        self.search_results_label = QLabel("")
        self.search_results_label.setStyleSheet("color: #1976D2; font-weight: bold; padding: 5px;")
        search_layout.addWidget(self.search_results_label)
        
        self.search_group.setLayout(search_layout)
        self.search_group.setVisible(False)  # Hidden by default
        main_layout.addWidget(self.search_group)
        
        # Selection toggle button
        self.selection_toggle_btn = QPushButton("‚úÖ Show Selection Tools")
        self.selection_toggle_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        self.selection_toggle_btn.clicked.connect(self.toggle_selection_panel)
        main_layout.addWidget(self.selection_toggle_btn)
        
        # Selection Controls Panel
        self.selection_group = QGroupBox("")
        selection_layout = QHBoxLayout()
        
        # Selection controls (Select All button hidden for now)
        
        self.clear_selection_btn = QPushButton("‚òê Clear Selection")
        self.clear_selection_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #616161;
            }
        """)
        self.clear_selection_btn.clicked.connect(self.clear_selection)
        selection_layout.addWidget(self.clear_selection_btn)
        
        # Separator
        selection_layout.addWidget(QLabel("|"))
        
        # Selection info
        self.selection_label = QLabel("Selected: 0 transactions")
        self.selection_label.setStyleSheet("font-weight: bold; color: #1976D2; padding: 8px;")
        selection_layout.addWidget(self.selection_label)
        
        # Spacer
        selection_layout.addStretch()
        
        # Action buttons
        self.edit_selected_btn = QPushButton("‚úèÔ∏è Edit Selected")
        self.edit_selected_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.edit_selected_btn.setEnabled(False)
        self.edit_selected_btn.clicked.connect(self.edit_selected_transactions)
        selection_layout.addWidget(self.edit_selected_btn)
        
        self.duplicate_selected_btn = QPushButton("üìã Duplicate Selected")
        self.duplicate_selected_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.duplicate_selected_btn.setEnabled(False)
        self.duplicate_selected_btn.clicked.connect(self.duplicate_selected_transactions)
        selection_layout.addWidget(self.duplicate_selected_btn)
        
        self.delete_selected_btn = QPushButton("üóëÔ∏è Delete Selected")
        self.delete_selected_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.delete_selected_btn.setEnabled(False)
        self.delete_selected_btn.clicked.connect(self.delete_selected_transactions)
        selection_layout.addWidget(self.delete_selected_btn)
        
        self.selection_group.setLayout(selection_layout)
        self.selection_group.setVisible(False)  # Hidden by default
        main_layout.addWidget(self.selection_group)
        
        # Table for displaying categorized data
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "‚òê", "Date", "Name", "Amount (‚Çπ)", "Category", "Description", "Reference"
        ])
        
        # Set table properties
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)             # Checkbox column - Fixed width
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Date
        header.setSectionResizeMode(2, QHeaderView.Stretch)           # Name
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Amount
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Category
        header.setSectionResizeMode(5, QHeaderView.Stretch)           # Description
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Reference
        
        # Set fixed width for checkbox column (50 pixels should be enough)
        self.table.setColumnWidth(0, 50)
        
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        main_layout.addWidget(self.table)
        
        # Bottom row: Total amount
        bottom_layout = QHBoxLayout()
        bottom_layout.addStretch()
        
        # Total amount label (centered)
        self.total_label = QLabel("Total: ‚Çπ0")
        total_font = QFont()
        total_font.setPointSize(14)
        total_font.setBold(True)
        self.total_label.setFont(total_font)
        self.total_label.setStyleSheet("""
            padding: 15px 30px; 
            background-color: #fff9c4; 
            border-radius: 5px; 
            border: 2px solid #fbc02d; 
            color: #f57f17;
        """)
        bottom_layout.addWidget(self.total_label)
        bottom_layout.addStretch()
        
        main_layout.addLayout(bottom_layout)
    
    # Menu bar and related code removed as per user request
    
    def manage_categories(self):
        """Open the category management dialog"""
        dialog = CategoryManagerDialog(self.categories, self)
        if dialog.exec() == QDialog.Accepted:
            # Get updated categories
            new_categories = dialog.get_categories()
            
            # Save to file
            if self.save_categories(new_categories):
                # Update in memory
                old_categories = self.categories.copy()
                self.categories = new_categories
                
                # Rebuild categorized_data structure
                self.rebuild_categorized_data()
                
                # Rebuild summary UI with new categories
                self.rebuild_summary_ui()
                
                # Re-categorize existing transactions if any
                if self.transactions:
                    self.recategorize_transactions()
                else:
                    # If no transactions, just update the summary display
                    self.update_summary()
                
                # Update display
                self.update_display()
                
                QMessageBox.information(
                    self, 
                    "Success", 
                    f"Categories updated successfully!\n\n"
                    f"Previous: {len(old_categories)} categories\n"
                    f"Current: {len(new_categories)} categories"
                )
    
    def rebuild_categorized_data(self):
        """Rebuild the categorized_data structure when categories change"""
        old_data = self.categorized_data.copy()
        
        # Create new structure
        self.categorized_data = {}
        for cat in self.categories:
            category_key = self.get_category_key(cat['name'], cat['fee'])
            self.categorized_data[category_key] = []
        
        # Clear transactions from old categories
        for transaction in self.transactions:
            transaction['category'] = None  # Reset category
    
    def rebuild_summary_ui(self):
        """Rebuild the summary UI when categories change"""
        # Clear existing labels
        for i in reversed(range(self.summary_layout.count())):
            widget = self.summary_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        # Clear the dictionary
        self.category_labels.clear()
        
        # Recreate labels based on new categories
        summary_font = QFont()
        summary_font.setPointSize(11)
        summary_font.setBold(True)
        
        # Calculate how many items per row (4 per row for good spacing)
        items_per_row = 4
        row = 0
        col = 0
        
        for cat in self.categories:
            category_key = self.get_category_key(cat['name'], cat['fee'])
            label = QLabel(f"{cat['name']}: 0 entries (‚Çπ0)")
            label.setFont(summary_font)
            label.setStyleSheet("padding: 10px; background-color: #e3f2fd; border-radius: 5px; color: #1a237e;")
            self.category_labels[category_key] = label
            
            # Add to grid layout
            self.summary_layout.addWidget(label, row, col)
            col += 1
            if col >= items_per_row:
                col = 0
                row += 1
                if row >= 2:  # Maximum 2 rows
                    break
    
    def recategorize_transactions(self):
        """Re-categorize all transactions based on updated categories"""
        if not self.transactions:
            return
        
        # Clear all categorized data
        for category_key in self.categorized_data:
            self.categorized_data[category_key] = []
        
        # Re-categorize each transaction
        for transaction in self.transactions:
            new_category = self.categorize_transaction(transaction['amount'])
            transaction['category'] = new_category
            self.categorized_data[new_category].append(transaction)
        
        QMessageBox.information(
            self,
            "Re-categorization Complete",
            f"Successfully re-categorized {len(self.transactions)} transactions based on updated categories."
        )
    
    def refresh_categories(self):
        """Reload categories from file"""
        try:
            old_count = len(self.categories)
            self.categories = self.load_categories()
            new_count = len(self.categories)
            
            # Rebuild structure
            self.rebuild_categorized_data()
            
            # Re-categorize if needed
            if self.transactions:
                self.recategorize_transactions()
            
            # Update display
            self.update_display()
            
            QMessageBox.information(
                self,
                "Categories Refreshed",
                f"Categories reloaded from file.\n\nCategories: {new_count} (was {old_count})"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to refresh categories: {e}")
    
    def show_about(self):
        """Show about dialog"""
        QMessageBox.about(
            self,
            "About Music Class Fee Tracker",
            """
            <h3>üéµ Music Class Fee Tracker</h3>
            <p><b>Version:</b> 2.0</p>
            <p><b>Author:</b> Your Name</p>
            <br>
            <p>A comprehensive tool for tracking and categorizing music class fees from bank statements.</p>
            <br>
            <p><b>Features:</b></p>
            <ul>
                <li>üì§ CSV file processing</li>
                <li>üîç Smart transaction categorization</li>
                <li>üìÇ Dynamic category management</li>
                <li>üîç Advanced search and filtering</li>
                <li>‚úèÔ∏è Transaction editing</li>
                <li>üíæ Export functionality</li>
            </ul>
            """
        )
        
    def upload_file(self):
        """Handle file upload and processing"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Bank Statement CSV",
            str(Path.home()),
            "CSV Files (*.csv *.CSV);;All Files (*)"
        )
        
        if file_path:
            try:
                loaded_from_save = self.process_csv(file_path)
                self.file_label.setText(f"File: {Path(file_path).name}")
                # Don't show success message here if loaded from save (already shown in process_csv)
                if not loaded_from_save:
                    QMessageBox.information(self, "Success", f"File processed successfully!\nLoaded {len(self.transactions)} transactions.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to process file:\n{str(e)}")
    
    def process_csv(self, file_path):
        """Process the CSV file and categorize transactions
        Returns True if loaded from saved data, False if parsed from CSV"""
        # Store current file path
        self.current_csv_file = file_path
        self.auto_save_file = self.get_auto_save_path(file_path)
        
        # Check if we have saved edits for this file
        saved_data = self.load_saved_data(file_path)
        
        if saved_data:
            # Ask user if they want to load saved edits
            reply = QMessageBox.question(
                self,
                "Load Saved Edits?",
                f"Found previously saved edits for this file.\n"
                f"Last saved: {saved_data.get('saved_at', 'Unknown')}\n\n"
                f"Would you like to load your saved edits?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                # Load saved transactions
                self.transactions = saved_data['transactions']
                
                # Rebuild categorized_data
                self.categorized_data = {}
                for cat in self.categories:
                    category_key = self.get_category_key(cat['name'], cat['fee'])
                    self.categorized_data[category_key] = []
                
                # Re-categorize loaded transactions
                for transaction in self.transactions:
                    category = transaction['category']
                    if category in self.categorized_data:
                        self.categorized_data[category].append(transaction)
                    else:
                        # If category doesn't exist, add it (handles old saved data with deleted categories)
                        self.categorized_data[category] = [transaction]
                
                # Update display with loaded data
                self.update_display()
                
                # Show success message with count
                QMessageBox.information(
                    self,
                    "Loaded Successfully",
                    f"Loaded {len(self.transactions)} saved transactions from previous edits."
                )
                return True  # Indicate we loaded from saved data
        
        # No saved data or user chose not to load - parse CSV fresh
        self.transactions = []
        # Rebuild categorized_data dynamically from categories
        self.categorized_data = {}
        for cat in self.categories:
            category_key = self.get_category_key(cat['name'], cat['fee'])
            self.categorized_data[category_key] = []
        
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
        # Find where the transaction data starts
        start_idx = 0
        for i, line in enumerate(lines):
            if 'Txn Date' in line and 'Credit' in line:
                start_idx = i + 1
                break
        
        # Process transactions
        for line in lines[start_idx:]:
            if not line.strip() or line.strip().startswith(','):
                continue
                
            # Parse the CSV line
            parts = self.parse_csv_line(line)
            
            if len(parts) >= 8:
                try:
                    # Extract transaction details
                    txn_date = parts[0].strip().strip('="').strip('"')
                    description = parts[3].strip().strip('="').strip('"')
                    credit = parts[6].strip().strip('="').strip('"').replace(',', '')
                    
                    if not credit or credit == '':
                        continue
                    
                    # Convert credit to float
                    amount = float(credit)
                    
                    # Extract short name from description (format: UPI/CR/xxx/NAME/...)
                    short_name = self.extract_name(description)
                    
                    # Get full name from mapping
                    full_name = get_full_name(short_name)
                    
                    # Categorize based on amount
                    category = self.categorize_transaction(amount)
                    
                    transaction = {
                        'date': txn_date,
                        'name': full_name,  # Use full name
                        'short_name': short_name,  # Keep short name for reference
                        'amount': amount,
                        'category': category,
                        'description': description,
                        'reference': parts[2].strip().strip('="').strip('"')
                    }
                    
                    self.transactions.append(transaction)
                    self.categorized_data[category].append(transaction)
                    
                except (ValueError, IndexError) as e:
                    # Skip invalid lines
                    continue
        
        self.update_display()
        return False  # Indicate we parsed from CSV
    
    def parse_csv_line(self, line):
        """Parse a CSV line handling the special format with = signs"""
        # This CSV uses ="" format for escaping, but also has regular "quoted" fields
        parts = []
        current = ""
        in_quotes = False
        
        i = 0
        while i < len(line):
            char = line[i]
            
            if char == '=' and i + 1 < len(line) and line[i + 1] == '"':
                # Start of a quoted field with ="" format
                i += 2
                field = ""
                while i < len(line):
                    if line[i] == '"' and (i + 1 >= len(line) or line[i + 1] in [',', '\n', '\r']):
                        break
                    field += line[i]
                    i += 1
                parts.append(field)
                i += 1  # Skip the closing quote
                # Skip the comma if present
                if i < len(line) and line[i] == ',':
                    i += 1
            elif char == '"' and not in_quotes:
                # Start of a regular quoted field
                in_quotes = True
                i += 1
            elif char == '"' and in_quotes:
                # End of a regular quoted field
                in_quotes = False
                parts.append(current)
                current = ""
                i += 1
                # Skip the comma if present
                if i < len(line) and line[i] == ',':
                    i += 1
            elif char == ',':
                if not in_quotes:
                    # Field separator (only when not in quotes)
                    parts.append(current)
                    current = ""
                    i += 1
                else:
                    # Comma inside quotes - part of the field value
                    current += char
                    i += 1
            else:
                current += char
                i += 1
        
        # Append the last field if any
        if current or in_quotes:
            parts.append(current)
        
        return parts
    
    def extract_name(self, description):
        """Extract name from UPI description"""
        # Format: UPI/CR/xxxxx/NAME/BANK/...
        parts = description.split('/')
        if len(parts) >= 4:
            name = parts[3].strip()
            # Clean up name
            name = re.sub(r'\s+', ' ', name)
            return name
        return "Unknown"
    
    def categorize_transaction(self, amount):
        """Categorize transaction based on amount using JSON config"""
        # Try to match with a fixed fee category
        for cat in self.categories:
            if cat['fee'] is not None and amount == cat['fee']:
                return self.get_category_key(cat['name'], cat['fee'])
        
        # If no match, use the first category with fee = null (variable amount)
        for cat in self.categories:
            if cat['fee'] is None:
                return self.get_category_key(cat['name'], cat['fee'])
        
        # Fallback (shouldn't happen if Donations exists)
        return "Uncategorized"
    
    def get_display_category(self, category):
        """Get category name without amount for display"""
        # Remove amount in brackets for cleaner display
        if '(‚Çπ' in category:
            return category.split(' (‚Çπ')[0]
        return category
    
    def update_display(self):
        """Update the table and summary with categorized data"""
        # Prevent concurrent updates
        if self._updating_display:
            return
        
        self._updating_display = True
        
        try:
            # Clear color cache to regenerate for any category changes
            self.color_cache.clear()
            
            # Clear selection since row indices will change
            self.selected_transactions.clear()
            self._manual_selection_panel_open = False
            
            # Clear table
            self.table.setRowCount(0)
            
            # Generate dynamic color mapping for all categories
            color_palette = [
                (QColor(200, 230, 201), QColor(27, 94, 32)),   # Light green / Dark green
                (QColor(187, 222, 251), QColor(13, 71, 161)),  # Light blue / Dark blue
                (QColor(255, 224, 178), QColor(230, 81, 0)),   # Light orange / Dark orange
                (QColor(248, 187, 208), QColor(136, 14, 79)),  # Light pink / Dark pink
                (QColor(225, 190, 231), QColor(74, 20, 140)),  # Light purple / Dark purple
                (QColor(255, 245, 157), QColor(245, 127, 23)), # Light yellow / Dark yellow
            ]
            
            colors = {}
            text_colors = {}
            for idx, category_key in enumerate(self.categorized_data.keys()):
                bg_color, fg_color = color_palette[idx % len(color_palette)]
                colors[category_key] = bg_color
                text_colors[category_key] = fg_color
            
            # Populate table by category (iterate dynamically)
            row = 0
            for category_key in self.categorized_data.keys():
                for transaction in self.categorized_data[category_key]:
                    self.table.insertRow(row)
                    
                    # Add checkbox for selection
                    checkbox_widget = QWidget()
                    checkbox_layout = QHBoxLayout(checkbox_widget)
                    checkbox = QCheckBox()
                    transaction_id = self.get_transaction_id(transaction)
                    checkbox.setChecked(transaction_id in self.selected_transactions)
                    checkbox.toggled.connect(lambda checked, tid=transaction_id: self.on_checkbox_toggled(tid, checked))
                    checkbox_layout.addWidget(checkbox)
                    checkbox_layout.setAlignment(Qt.AlignCenter)
                    checkbox_layout.setContentsMargins(0, 0, 0, 0)
                    self.table.setCellWidget(row, 0, checkbox_widget)
                
                    # Add data to cells (shifted by 1 column)
                    date_item = QTableWidgetItem(transaction['date'])
                    date_item.setFlags(date_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, 1, date_item)
                    
                    name_item = QTableWidgetItem(transaction['name'])
                    name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, 2, name_item)
                    
                    amount_item = QTableWidgetItem(f"‚Çπ{transaction['amount']:.2f}")
                    amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    amount_item.setFlags(amount_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, 3, amount_item)
                    
                    # Display category without amount
                    display_category = self.get_display_category(transaction['category'])
                    category_item = QTableWidgetItem(display_category)
                    category_item.setBackground(colors[transaction['category']])
                    category_item.setForeground(text_colors[transaction['category']])
                    category_item.setTextAlignment(Qt.AlignCenter)
                    category_item.setFont(QFont("", -1, QFont.Bold))
                    category_item.setFlags(category_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, 4, category_item)
                    
                    desc_item = QTableWidgetItem(transaction['description'][:50] + "..." 
                                                if len(transaction['description']) > 50 
                                                else transaction['description'])
                    desc_item.setFlags(desc_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, 5, desc_item)
                    
                    ref_item = QTableWidgetItem(transaction['reference'])
                    ref_item.setFlags(ref_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, 6, ref_item)
                    
                    row += 1
        
            # Update summary and selection UI
            self.update_summary()
            self.update_selection_ui()
            
        finally:
            self._updating_display = False
    
    def update_summary(self):
        """Update summary statistics dynamically"""
        # Color palette for categories
        colors = ['#c8e6c9', '#bbdefb', '#ffe0b2', '#f8bbd0', '#e1bee7', '#fff9c4']
        text_colors = ['#1b5e20', '#0d47a1', '#e65100', '#880e4f', '#4a148c', '#f57f17']
        
        total_amount = 0
        fees_only_amount = 0  # Total excluding donations
        
        # Update each category label dynamically
        for idx, (category_key, label) in enumerate(self.category_labels.items()):
            count = len(self.categorized_data.get(category_key, []))
            cat_total = sum(t['amount'] for t in self.categorized_data.get(category_key, []))
            total_amount += cat_total
            
            # Get category name (without amount)
            cat_name = category_key.split(' (‚Çπ')[0] if '(‚Çπ' in category_key else category_key
            
            # Add to fees total only if it's not a donation/uncategorised category
            # Check if category has a fixed fee (not None)
            category_obj = next((cat for cat in self.categories if self.get_category_key(cat['name'], cat['fee']) == category_key), None)
            if category_obj and category_obj['fee'] is not None:
                fees_only_amount += cat_total
            
            label.setText(f"{cat_name}: {count} entries (‚Çπ{cat_total:.2f})")
            
            # Update colors based on whether there are entries
            bg_color = colors[idx % len(colors)] if count > 0 else '#e3f2fd'
            fg_color = text_colors[idx % len(text_colors)] if count > 0 else '#1a237e'
            
            label.setStyleSheet(
                f"padding: 12px; background-color: {bg_color}; border-radius: 5px; color: {fg_color};"
            )
        
        # Update totals - show both fees and grand total
        self.total_label.setText(f"Fee Collections: ‚Çπ{fees_only_amount:.2f}  |  Grand Total: ‚Çπ{total_amount:.2f}")
        self.total_label.setStyleSheet(
            "padding: 10px; background-color: #fff9c4; border-radius: 5px; border: 2px solid #fbc02d; color: #f57f17;"
        )
    
    def on_search_text_changed(self):
        """Handle search text changes with debouncing"""
        self.search_timer.stop()  # Stop any existing timer
        self.search_timer.start(self.search_delay)  # Start new timer
    
    def perform_search(self):
        """Optimized search with caching and efficient filtering"""
        search_term = self.search_input.text().strip().lower()
        search_field = self.search_field_combo.currentText()
        
        if not search_term:
            # If search is empty, show all transactions
            self.update_display()
            self.search_results_label.setText("")
            return
        
        # Cache color mapping to avoid regeneration
        if not self.color_cache:
            self._generate_color_cache()
        
        # Fast filtering using list comprehension
        filtered_transactions = self._filter_transactions_fast(search_term, search_field)
        
        # Efficient table update
        self._update_table_efficiently(filtered_transactions)
        
        # Update search results label
        total = len(self.transactions)
        found = len(filtered_transactions)
        self.search_results_label.setText(f"Found {found} of {total} transactions")
    
    def _generate_color_cache(self):
        """Generate and cache color mappings once"""
        color_palette = [
            (QColor(200, 230, 201), QColor(27, 94, 32)),   # Light green / Dark green
            (QColor(187, 222, 251), QColor(13, 71, 161)),  # Light blue / Dark blue
            (QColor(255, 224, 178), QColor(230, 81, 0)),   # Light orange / Dark orange
            (QColor(248, 187, 208), QColor(136, 14, 79)),  # Light pink / Dark pink
            (QColor(225, 190, 231), QColor(74, 20, 140)),  # Light purple / Dark purple
            (QColor(255, 245, 157), QColor(245, 127, 23)), # Light yellow / Dark yellow
        ]
        
        self.color_cache = {'colors': {}, 'text_colors': {}}
        for idx, category_key in enumerate(self.categorized_data.keys()):
            bg_color, fg_color = color_palette[idx % len(color_palette)]
            self.color_cache['colors'][category_key] = bg_color
            self.color_cache['text_colors'][category_key] = fg_color
    
    def _filter_transactions_fast(self, search_term, search_field):
        """Fast transaction filtering using optimized logic"""
        if search_field == "All":
            # For "All" search, do exact amount matching but substring for text fields
            return [t for t in self.transactions if (
                search_term in t['name'].lower() or
                search_term in t['category'].lower() or
                search_term in t['description'].lower() or
                self._amount_matches(search_term, t['amount'])
            )]
        elif search_field == "Name":
            return [t for t in self.transactions if search_term in t['name'].lower()]
        elif search_field == "Amount":
            # For amount-specific search, use smart matching
            return [t for t in self.transactions if self._amount_matches(search_term, t['amount'])]
        elif search_field == "Category":
            return [t for t in self.transactions if search_term in t['category'].lower()]
        elif search_field == "Description":
            return [t for t in self.transactions if search_term in t['description'].lower()]
        return []
    
    def _amount_matches(self, search_term, amount):
        """Smart amount matching that handles both exact and partial matches intelligently"""
        formatted_amount = f"{amount:.2f}"
        
        # If search term looks like a complete amount (has decimal point), do exact match
        if '.' in search_term:
            return formatted_amount == search_term
        
        # Otherwise, check if it matches the whole number part exactly
        # This prevents "2" from matching "502" but allows "2" to match "2.00"
        try:
            # Try to parse as a number
            search_num = float(search_term)
            return amount == search_num or formatted_amount.startswith(search_term + '.')
        except ValueError:
            # If not a valid number, fall back to substring match
            return search_term in formatted_amount
    
    def _update_table_efficiently(self, filtered_transactions):
        """Efficiently update table without recreating widgets unnecessarily"""
        # Clear table contents but preserve structure
        self.table.setRowCount(len(filtered_transactions))
        
        colors = self.color_cache['colors']
        text_colors = self.color_cache['text_colors']
        
        # Batch table updates for better performance
        for row, transaction in enumerate(filtered_transactions):
            # Add checkbox for selection  
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox = QCheckBox()
            transaction_id = self.get_transaction_id(transaction)
            checkbox.setChecked(transaction_id in self.selected_transactions)
            checkbox.toggled.connect(lambda checked, tid=transaction_id: self.on_checkbox_toggled(tid, checked))
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 0, checkbox_widget)
            
            # Set basic data (shifted by 1 column)
            date_item = QTableWidgetItem(transaction['date'])
            date_item.setFlags(date_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 1, date_item)
            
            name_item = QTableWidgetItem(transaction['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 2, name_item)
            
            # Amount with alignment
            amount_item = QTableWidgetItem(f"‚Çπ{transaction['amount']:.2f}")
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            amount_item.setFlags(amount_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 3, amount_item)
            
            # Category with cached colors
            display_category = self.get_display_category(transaction['category'])
            category_item = QTableWidgetItem(display_category)
            category_item.setBackground(colors[transaction['category']])
            category_item.setForeground(text_colors[transaction['category']])
            category_item.setTextAlignment(Qt.AlignCenter)
            category_item.setFont(QFont("", -1, QFont.Bold))
            category_item.setFlags(category_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 4, category_item)
            
            # Description (truncated)
            desc_text = (transaction['description'][:50] + "..." 
                        if len(transaction['description']) > 50 
                        else transaction['description'])
            desc_item = QTableWidgetItem(desc_text)
            desc_item.setFlags(desc_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 5, desc_item)
            
            ref_item = QTableWidgetItem(transaction['reference'])
            ref_item.setFlags(ref_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 6, ref_item)
    

    def on_checkbox_toggled(self, transaction_id, checked):
        """Handle checkbox toggle (most reliable signal)"""
        if checked:
            self.selected_transactions.add(transaction_id)
        else:
            self.selected_transactions.discard(transaction_id)
        
        self.update_selection_ui()
        
    def on_checkbox_state_changed(self, checkbox, transaction_id):
        """Handle checkbox state changes (simplified)"""
        is_checked = checkbox.isChecked()
        
        if is_checked:
            self.selected_transactions.add(transaction_id)
        else:
            self.selected_transactions.discard(transaction_id)
        
        self.update_selection_ui()
    
    def on_checkbox_changed(self, transaction_id, transaction, state):
        """Handle checkbox state changes (legacy - keeping for compatibility)"""
        if state == Qt.Checked:
            self.selected_transactions.add(transaction_id)
        else:
            self.selected_transactions.discard(transaction_id)
        
        self.update_selection_ui()
    
    def get_visible_transactions(self):
        """Get currently visible transactions in table order"""
        visible = []
        for row in range(self.table.rowCount()):
            # Get transaction data from table
            name = self.table.item(row, 2).text() if self.table.item(row, 2) else ""
            amount_text = self.table.item(row, 3).text() if self.table.item(row, 3) else ""
            
            # Find matching transaction
            for transaction in self.transactions:
                if (transaction['name'] == name and 
                    f"‚Çπ{transaction['amount']:.2f}" == amount_text):
                    visible.append(transaction)
                    break
        return visible
    
    def select_all_transactions(self):
        """Select all visible transactions"""
        visible_transactions = self.get_visible_transactions()
        
        for transaction in visible_transactions:
            transaction_id = self.get_transaction_id(transaction)
            self.selected_transactions.add(transaction_id)
        
        # Update all checkboxes
        for row in range(self.table.rowCount()):
            checkbox_widget = self.table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)
        
        self.update_selection_ui()
    
    def clear_selection(self):
        """Clear all selections"""
        self.selected_transactions.clear()
        
        # Update all checkboxes
        for row in range(self.table.rowCount()):
            checkbox_widget = self.table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(False)
        
        # Reset manual flag when clearing
        self._manual_selection_panel_open = False
        self.update_selection_ui()
    
    def update_selection_ui(self):
        """Update selection-related UI elements"""
        count = len(self.selected_transactions)
        
        self.selection_label.setText(f"Selected: {count} transactions")
        
        # Enable/disable action buttons based on selection
        has_selection = count > 0
        self.edit_selected_btn.setEnabled(has_selection)
        self.duplicate_selected_btn.setEnabled(has_selection)
        self.delete_selected_btn.setEnabled(has_selection)
        
        # Select all button is hidden for now - no updates needed
        
        # Smart accordion behavior
        self.auto_toggle_selection_panel(has_selection)
    
    def toggle_selection_panel(self):
        """Manually toggle the selection panel visibility"""
        if self.selection_group.isVisible():
            self.hide_selection_panel_manually()
        else:
            self.show_selection_panel_manually()
    
    def auto_toggle_selection_panel(self, has_selection):
        """Automatically show/hide selection panel based on selections"""
        if has_selection and not self.selection_group.isVisible():
            # Auto-show when selections are made
            self.selection_group.setVisible(True)
            self.selection_toggle_btn.setText("‚ùå Hide Selection Tools")
        elif not has_selection and self.selection_group.isVisible():
            # Auto-hide when no selections, but respect manual override
            if not self._manual_selection_panel_open:
                self.selection_group.setVisible(False)
                self.selection_toggle_btn.setText("‚úÖ Show Selection Tools")
        
        # Reset manual flag when selections exist (user is actively using selection)
        if has_selection:
            self._manual_selection_panel_open = False
    
    def show_selection_panel_manually(self):
        """Show selection panel and mark as manually opened"""
        self._manual_selection_panel_open = True
        self.selection_group.setVisible(True)
        self.selection_toggle_btn.setText("‚ùå Hide Selection Tools")
    
    def hide_selection_panel_manually(self):
        """Hide selection panel and clear manual flag"""
        self._manual_selection_panel_open = False
        self.selection_group.setVisible(False)
        self.selection_toggle_btn.setText("‚úÖ Show Selection Tools")
    
    def get_selected_transactions(self):
        """Get the actual transaction objects for selected transactions"""
        selected = []
        for transaction in self.transactions:
            transaction_id = self.get_transaction_id(transaction)
            if transaction_id in self.selected_transactions:
                selected.append(transaction)
        return selected
    
    def edit_selected_transactions(self):
        """Edit selected transactions"""
        selected = self.get_selected_transactions()
        if not selected:
            QMessageBox.information(self, "No Selection", "Please select transactions to edit.")
            return
        
        if len(selected) == 1:
            # Edit single transaction
            self.edit_transaction(selected[0])
        else:
            # Bulk edit - show dialog for category change
            self.bulk_edit_transactions(selected)
    
    def bulk_edit_transactions(self, transactions):
        """Show bulk edit dialog for multiple transactions"""
        from PySide6.QtWidgets import QInputDialog
        
        # Get available categories
        categories = list(self.categorized_data.keys())
        
        # Show category selection dialog
        category, ok = QInputDialog.getItem(
            self, 
            "Bulk Edit", 
            f"Change category for {len(transactions)} transactions:",
            categories, 
            0, 
            False
        )
        
        if ok and category:
            # Apply category change to all selected transactions
            for transaction in transactions:
                old_category = transaction['category']
                if old_category != category:
                    # Remove from old category
                    if transaction in self.categorized_data[old_category]:
                        self.categorized_data[old_category].remove(transaction)
                    # Add to new category
                    self.categorized_data[category].append(transaction)
                    # Update transaction
                    transaction['category'] = category
            
            # Refresh display
            # If search is active, re-run search to preserve filtered view
            if self.search_input.text().strip():
                self.perform_search()
            else:
                self.update_display()
            
            # Auto-save changes
            self.auto_save_data()
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Updated {len(transactions)} transactions to {category.split(' (‚Çπ')[0]}"
            )
    
    def delete_selected_transactions(self):
        """Delete selected transactions with confirmation"""
        selected = self.get_selected_transactions()
        if not selected:
            QMessageBox.information(self, "No Selection", "Please select transactions to delete.")
            return
        
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Confirm Deletion", 
            f"Are you sure you want to delete {len(selected)} selected transactions?\n\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Remove from categorized data and main transactions list
            for transaction in selected:
                # Remove from selection tracking
                transaction_id = self.get_transaction_id(transaction)
                self.selected_transactions.discard(transaction_id)
                
                # Remove from category
                category = transaction['category']
                if transaction in self.categorized_data[category]:
                    self.categorized_data[category].remove(transaction)
                # Remove from main list
                if transaction in self.transactions:
                    self.transactions.remove(transaction)
            
            # Clear selection and refresh display
            self.clear_selection()
            # If search is active, re-run search to preserve filtered view
            if self.search_input.text().strip():
                self.perform_search()
            else:
                self.update_display()
            
            # Auto-save changes
            self.auto_save_data()
            
            QMessageBox.information(
                self, 
                "Deleted", 
                f"Successfully deleted {len(selected)} transactions."
            )
    
    def duplicate_selected_transactions(self):
        """Duplicate selected transactions"""
        selected = self.get_selected_transactions()
        if not selected:
            QMessageBox.information(self, "No Selection", "Please select transactions to duplicate.")
            return
        
        # Track how many were duplicated
        duplicated_count = 0
        
        for transaction in selected:
            # Open edit dialog with duplicated transaction
            # Make a copy of the transaction
            duplicated_transaction = {
                'date': transaction['date'],
                'name': transaction['name'],
                'short_name': transaction.get('short_name', transaction['name']),
                'amount': transaction['amount'],
                'category': transaction['category'],
                'description': transaction['description'],
                'reference': transaction['reference'] + '_COPY'  # Mark as copy
            }
            
            # Get available categories for the dialog
            categories = list(self.categorized_data.keys())
            
            # Open edit dialog
            dialog = EditTransactionDialog(duplicated_transaction, categories, self)
            
            if dialog.exec() == QDialog.Accepted:
                # Get the edited duplicate
                new_transaction = dialog.get_updated_transaction()
                
                # Add to transactions list
                self.transactions.append(new_transaction)
                
                # Add to appropriate category
                category = new_transaction['category']
                if category in self.categorized_data:
                    self.categorized_data[category].append(new_transaction)
                
                duplicated_count += 1
        
        if duplicated_count > 0:
            # Clear selection
            self.selected_transactions.clear()
            
            # Refresh display
            # If search is active, re-run search to preserve filtered view
            if self.search_input.text().strip():
                self.perform_search()
            else:
                self.update_display()
            
            # Auto-save changes
            self.auto_save_data()
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Successfully duplicated {duplicated_count} transaction(s)."
            )

    def set_search_delay(self, delay_ms):
        """Adjust search delay (useful for performance tuning)"""
        self.search_delay = delay_ms
    
    def toggle_search(self):
        """Toggle the visibility of the search section"""
        if self.search_group.isVisible():
            # Stop any pending search operations immediately
            self.search_timer.stop()
            
            self.search_group.setVisible(False)
            self.search_toggle_btn.setText("üîç Show Search")
            # Clear search when hiding
            self.clear_search()
        else:
            self.search_group.setVisible(True)
            self.search_toggle_btn.setText("üîç Hide Search")
            # Focus on search input when showing
            self.search_input.setFocus()
    
    def clear_search(self):
        """Clear search and show all transactions"""
        # Stop any pending search operations
        self.search_timer.stop()
        
        # Check if there was actually a search term to clear
        had_search_term = bool(self.search_input.text().strip())
        
        self.search_input.clear()
        self.search_field_combo.setCurrentIndex(0)
        self.search_results_label.setText("")
        
        # Only update display if there was a search active (optimization)
        if had_search_term:
            self.update_display()
    
    def edit_transaction(self, transaction):
        """Edit a transaction"""
        
        # Get list of available categories for the dropdown
        available_categories = list(self.categorized_data.keys())
        dialog = EditTransactionDialog(transaction, available_categories, self)
        if dialog.exec() == QDialog.Accepted:
            # Get all updated values from the dialog
            updated_transaction = dialog.get_updated_transaction()
                        
            old_category = transaction['category']
            new_category = updated_transaction['category']
            
            # Update all transaction fields (this updates both self.transactions and categorized_data since they're references)
            transaction['name'] = updated_transaction['name']
            transaction['date'] = updated_transaction['date']
            transaction['amount'] = updated_transaction['amount']
            transaction['category'] = new_category  # Always update category
                        
            # Handle category change if necessary
            if old_category != new_category:
                # Remove from old category
                self.categorized_data[old_category].remove(transaction)
                # Add to new category
                self.categorized_data[new_category].append(transaction)
            
            # Refresh display to show changes
            # If search is active, re-run search to preserve filtered view
            if self.search_input.text().strip():
                self.perform_search()
            else:
                self.update_display()
            
            # Auto-save changes
            self.auto_save_data()
            
            # Show confirmation message
            QMessageBox.information(
                self, 
                "Success", 
                f"Transaction updated successfully!\n\nName: {transaction['name']}\nDate: {transaction['date']}\nAmount: ‚Çπ{transaction['amount']:.2f}\nCategory: {transaction['category']}"
            )
    
    def delete_transaction(self, transaction):
        """Delete a transaction"""
        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete this transaction?\n\n"
            f"Name: {transaction['name']}\n"
            f"Amount: ‚Çπ{transaction['amount']:.2f}\n"
            f"Date: {transaction['date']}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Remove from main transactions list (more efficient)
            self.transactions.remove(transaction)
            
            # Remove from categorized data
            category = transaction['category']
            self.categorized_data[category].remove(transaction)
            
            # Refresh display
            self.update_display()
            
            # Auto-save changes
            self.auto_save_data()
    
    def export_summary(self):
        """Export summary to an Excel file with formatting"""
        if not self.transactions:
            QMessageBox.warning(self, "No Data", "Please upload a CSV file first.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Summary to Excel",
            str(Path.home() / f"fee_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                # Create workbook
                wb = Workbook()
                
                # === Summary Sheet ===
                ws_summary = wb.active
                ws_summary.title = "Summary"
                
                # Title
                ws_summary['A1'] = "MUSIC CLASS FEE TRACKER - SUMMARY REPORT"
                ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
                ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws_summary.merge_cells('A1:E1')
                ws_summary.row_dimensions[1].height = 30
                
                # Date
                ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_summary['A2'].font = Font(italic=True)
                ws_summary.merge_cells('A2:E2')
                
                # Headers
                row = 4
                headers = ['Category', 'Number of Entries', 'Total Amount (‚Çπ)', 'Average Amount (‚Çπ)', '% of Total']
                for col, header in enumerate(headers, 1):
                    cell = ws_summary.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Calculate grand total for percentages
                grand_total = sum(t['amount'] for t in self.transactions)
                
                # Data rows
                row = 5
                color_palette = ['D9E1F2', 'E2EFDA', 'FCE4D6', 'F4B084', 'C5E0B4', 'FFE699']
                
                for idx, category_key in enumerate(self.categorized_data.keys()):
                    transactions = self.categorized_data[category_key]
                    total = sum(t['amount'] for t in transactions)
                    count = len(transactions)
                    avg = total / count if count > 0 else 0
                    percentage = (total / grand_total) if grand_total > 0 else 0  # Store as decimal (0.2836), not percentage (28.36)
                    
                    # Category name (without amount suffix)
                    category_name = category_key.split(' (‚Çπ')[0] if '(‚Çπ' in category_key else category_key
                    
                    ws_summary.cell(row=row, column=1, value=category_name)
                    ws_summary.cell(row=row, column=2, value=count)
                    ws_summary.cell(row=row, column=3, value=total).number_format = '‚Çπ#,##0.00'
                    ws_summary.cell(row=row, column=4, value=avg).number_format = '‚Çπ#,##0.00'
                    ws_summary.cell(row=row, column=5, value=percentage).number_format = '0.00%'  # This format multiplies by 100
                    
                    # Apply color fill
                    fill_color = color_palette[idx % len(color_palette)]
                    for col in range(1, 6):
                        cell = ws_summary.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        if col > 1:
                            cell.alignment = Alignment(horizontal="right")
                    
                    row += 1
                
                # Grand Total Row
                ws_summary.cell(row=row, column=1, value="GRAND TOTAL")
                ws_summary.cell(row=row, column=1).font = Font(bold=True)
                ws_summary.cell(row=row, column=2, value=len(self.transactions))
                ws_summary.cell(row=row, column=3, value=grand_total).number_format = '‚Çπ#,##0.00'
                
                for col in range(1, 6):
                    cell = ws_summary.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium')
                    )
                    if col > 1:
                        cell.alignment = Alignment(horizontal="right")
                
                # Adjust column widths
                ws_summary.column_dimensions['A'].width = 30
                ws_summary.column_dimensions['B'].width = 18
                ws_summary.column_dimensions['C'].width = 18
                ws_summary.column_dimensions['D'].width = 20
                ws_summary.column_dimensions['E'].width = 15
                
                # === Detailed Transactions Sheet ===
                ws_detail = wb.create_sheet("All Transactions")
                
                # Headers
                detail_headers = ['Date', 'Name', 'Amount (‚Çπ)', 'Category', 'Description', 'Reference']
                for col, header in enumerate(detail_headers, 1):
                    cell = ws_detail.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Data rows
                row = 2
                for transaction in self.transactions:
                    ws_detail.cell(row=row, column=1, value=transaction['date'])
                    ws_detail.cell(row=row, column=2, value=transaction['name'])
                    ws_detail.cell(row=row, column=3, value=transaction['amount']).number_format = '‚Çπ#,##0.00'
                    
                    # Category name without amount
                    category_name = transaction['category'].split(' (‚Çπ')[0] if '(‚Çπ' in transaction['category'] else transaction['category']
                    ws_detail.cell(row=row, column=4, value=category_name)
                    ws_detail.cell(row=row, column=5, value=transaction['description'])
                    ws_detail.cell(row=row, column=6, value=transaction['reference'])
                    
                    # Apply borders and alignment
                    for col in range(1, 7):
                        cell = ws_detail.cell(row=row, column=col)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        if col == 3:  # Amount column
                            cell.alignment = Alignment(horizontal="right")
                    
                    row += 1
                
                # Adjust column widths
                ws_detail.column_dimensions['A'].width = 15
                ws_detail.column_dimensions['B'].width = 25
                ws_detail.column_dimensions['C'].width = 15
                ws_detail.column_dimensions['D'].width = 20
                ws_detail.column_dimensions['E'].width = 40
                ws_detail.column_dimensions['F'].width = 20
                
                # === Category-wise Sheets ===
                for category_key in self.categorized_data.keys():
                    transactions = self.categorized_data[category_key]
                    if not transactions:
                        continue
                    
                    # Create sheet name (max 31 chars, no special chars)
                    category_name = category_key.split(' (‚Çπ')[0] if '(‚Çπ' in category_key else category_key
                    sheet_name = category_name[:31].replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
                    
                    ws_cat = wb.create_sheet(sheet_name)
                    
                    # Title
                    ws_cat['A1'] = f"{category_name} - Detailed Report"
                    ws_cat['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                    ws_cat['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    ws_cat['A1'].alignment = Alignment(horizontal="center", vertical="center")
                    ws_cat.merge_cells('A1:F1')
                    ws_cat.row_dimensions[1].height = 25
                    
                    # Headers
                    for col, header in enumerate(detail_headers, 1):
                        cell = ws_cat.cell(row=3, column=col, value=header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    
                    # Data
                    row = 4
                    for transaction in transactions:
                        ws_cat.cell(row=row, column=1, value=transaction['date'])
                        ws_cat.cell(row=row, column=2, value=transaction['name'])
                        ws_cat.cell(row=row, column=3, value=transaction['amount']).number_format = '‚Çπ#,##0.00'
                        ws_cat.cell(row=row, column=4, value=category_name)
                        ws_cat.cell(row=row, column=5, value=transaction['description'])
                        ws_cat.cell(row=row, column=6, value=transaction['reference'])
                        
                        for col in range(1, 7):
                            cell = ws_cat.cell(row=row, column=col)
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            if col == 3:
                                cell.alignment = Alignment(horizontal="right")
                        
                        row += 1
                    
                    # Total row
                    total = sum(t['amount'] for t in transactions)
                    ws_cat.cell(row=row, column=2, value="TOTAL:")
                    ws_cat.cell(row=row, column=2).font = Font(bold=True)
                    ws_cat.cell(row=row, column=3, value=total).number_format = '‚Çπ#,##0.00'
                    ws_cat.cell(row=row, column=3).font = Font(bold=True)
                    ws_cat.cell(row=row, column=2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    ws_cat.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    
                    # Adjust column widths
                    ws_cat.column_dimensions['A'].width = 15
                    ws_cat.column_dimensions['B'].width = 25
                    ws_cat.column_dimensions['C'].width = 15
                    ws_cat.column_dimensions['D'].width = 20
                    ws_cat.column_dimensions['E'].width = 40
                    ws_cat.column_dimensions['F'].width = 20
                
                # Save workbook
                wb.save(file_path)
                
                QMessageBox.information(
                    self, 
                    "Success", 
                    f"Excel summary exported successfully!\n\n"
                    f"File: {Path(file_path).name}\n"
                    f"Location: {Path(file_path).parent}\n\n"
                    f"Sheets included:\n"
                    f"‚Ä¢ Summary (overview)\n"
                    f"‚Ä¢ All Transactions (complete list)\n"
                    f"‚Ä¢ Individual category sheets"
                )
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export Excel summary:\n{str(e)}")


def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    
    # Set application style
    app.setStyle("Fusion")
    
    window = FeeTrackerApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
